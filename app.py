"""
SE Grade Calculator - Flask Backend
Software Engineering Program

Theory Marks Distribution:
  Mid Term   : out of 30  →  30% weight
  Sessional  : out of 20  →  20% weight
  Final Term : out of 50  →  50% weight
  (or enter total directly out of 100)

Lab Marks Distribution:
  Lab marks  : out of 100 → 100% (1 credit hour)
  Lab is graded independently from theory.
"""

from flask import Flask, render_template, request, jsonify, send_file
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from datetime import datetime
import io
import json
import os
import base64
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)

# Path where marks are saved between sessions
SAVE_FILE = os.path.join(os.path.dirname(__file__), "saved_data.json")


# ─────────────────────────────────────────────────────────────────────────────
# COURSE DATABASE
# Add more semesters below following the same format.
# teacher_photo → place the image file in static/images/ with this filename.
# ─────────────────────────────────────────────────────────────────────────────
COURSE_DATABASE = {
    "Semester 1": [],
    "Semester 2": [],
    "Semester 3": [],
    "Semester 4": [],
    "Semester 5": [
        {
            "id":             "FM",
            "course_name":    "Formal Methods",
            "teacher":        "Sir Yasir",
            "theory_credits": 3,
            "has_lab":        False,
            "lab_credits":    0,
            "total_credits":  3,
            "teacher_photo":  "yasir.jpg"
        },
        {
            "id":             "IDS",
            "course_name":    "Introduction to Data Science",
            "teacher":        "",
            "theory_credits": 2,
            "has_lab":        True,
            "lab_credits":    1,
            "total_credits":  3,
            "teacher_photo":  "ids_teacher.jpg"
        },
        {
            "id":             "OS",
            "course_name":    "Operating Systems",
            "teacher":        "Saeed Akbar",
            "theory_credits": 3,
            "has_lab":        True,
            "lab_credits":    1,
            "total_credits":  4,
            "teacher_photo":  "saeed_akbar.jpg"
        },
        {
            "id":             "SDA",
            "course_name":    "Software Design & Architecture",
            "teacher":        "Mubashir Husain",
            "theory_credits": 3,
            "has_lab":        False,
            "lab_credits":    0,
            "total_credits":  3,
            "teacher_photo":  "mubashir.jpg"
        },
        {
            "id":             "WDD",
            "course_name":    "Web Design & Development",
            "teacher":        "Nasir Khan",
            "theory_credits": 2,
            "has_lab":        True,
            "lab_credits":    1,
            "total_credits":  3,
            "teacher_photo":  "nasir_khan.jpg"
        },
    ],
    "Semester 6": [],
    "Semester 7": [],
    "Semester 8": [],
}


# ─────────────────────────────────────────────────────────────────────────────
# GRADE CALCULATION HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def compute_total_percent(mid=0, sessional=0, final=0, out_of_100=False, total_100=0):
    """
    Breakdown mode : mid/30 × 30 + sessional/20 × 20 + final/50 × 50
    Out-of-100 mode: user supplies the total directly (clamped 0–100)
    """
    if out_of_100:
        return round(max(0.0, min(100.0, float(total_100))), 2)

    mid_part       = (float(mid)       / 30) * 30
    sessional_part = (float(sessional) / 20) * 20
    final_part     = (float(final)     / 50) * 50
    return round(mid_part + sessional_part + final_part, 2)


def grade_from_percent(pct):
    """Return (letter_grade, grade_point, remark) for a percentage 0–100."""
    if   pct >= 85: return "A",  4.00, "Excellent"
    elif pct >= 78: return "B+", 3.50, "Outstanding"
    elif pct >= 70: return "B",  3.00, "Good"
    elif pct >= 65: return "C+", 2.50, "Above Average"
    elif pct >= 60: return "C",  2.00, "Average"
    elif pct >= 55: return "D+", 1.50, "Below Average"
    elif pct >= 50: return "D",  1.00, "Poor but Passing"
    else:           return "F",  0.00, "Failing"


# ─────────────────────────────────────────────────────────────────────────────
# SAVE / LOAD  (persists marks + teacher photos to saved_data.json)
# ─────────────────────────────────────────────────────────────────────────────

def load_saved():
    """Load all saved course data from disk. Returns empty dict if file missing."""
    if os.path.exists(SAVE_FILE):
        with open(SAVE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def write_saved(data):
    """Write all saved course data to disk."""
    with open(SAVE_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# ─────────────────────────────────────────────────────────────────────────────
# ROUTES
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/semesters")
def get_semesters():
    """Return list of available semester names."""
    return jsonify(list(COURSE_DATABASE.keys()))


@app.route("/api/courses/<semester>")
def get_courses(semester):
    """Return all courses for a semester, merged with any saved marks."""
    courses = COURSE_DATABASE.get(semester, [])
    saved   = load_saved()
    result  = []
    for course in courses:
        key        = f"{semester}__{course['id']}"
        saved_data = saved.get(key, {})
        result.append({**course, **saved_data})   # saved values override defaults
    return jsonify(result)


@app.route("/api/save", methods=["POST"])
def save_course():
    """
    Save marks and optional teacher photo for one course.
    Expects JSON: { semester, course_id, midterm_marks, sessional_marks,
                    final_marks, out_of_100, total_100_marks,
                    photo_b64 (optional), photo_filename (optional) }
    """
    data      = request.get_json()
    semester  = data.get("semester", "")
    course_id = data.get("course_id", "")

    if not semester or not course_id:
        return jsonify({"error": "Missing semester or course_id"}), 400

    key   = f"{semester}__{course_id}"
    saved = load_saved()

    # Save teacher photo to static/images/ if one was provided as base64
    photo_b64  = data.get("photo_b64", "")
    photo_name = data.get("photo_filename", "")
    if photo_b64 and photo_name:
        img_dir  = os.path.join(os.path.dirname(__file__), "static", "images")
        os.makedirs(img_dir, exist_ok=True)
        img_path = os.path.join(img_dir, photo_name)
        encoded  = photo_b64.split(",", 1)[1] if "," in photo_b64 else photo_b64
        with open(img_path, "wb") as f:
            f.write(base64.b64decode(encoded))
        data["teacher_photo"] = photo_name   # store the filename, not the raw base64

    # Strip fields that don't belong in the JSON record
    skip  = {"semester", "course_id", "photo_b64", "photo_filename"}
    entry = {k: v for k, v in data.items() if k not in skip}
    saved[key] = entry
    write_saved(saved)
    return jsonify({"ok": True, "key": key})


@app.route("/api/calculate", methods=["POST"])
def calculate():
    """
    Calculate GPA and CGPA for a list of courses.
    Expects JSON: { courses: [...], previous_cgpa, previous_credits }
    """
    data         = request.get_json()
    courses      = data.get("courses", [])
    prev_cgpa    = float(data.get("previous_cgpa",    0))
    prev_credits = float(data.get("previous_credits", 0))

    if not courses:
        return jsonify({"error": "No courses provided"}), 400

    results      = []
    sem_gp       = 0.0
    sem_credits  = 0.0

    for c in courses:
        theory_cr = float(c.get("theory_credits", 3))
        has_lab   = bool(c.get("has_lab", False))
        lab_cr    = float(c.get("lab_credits", 0)) if has_lab else 0.0
        total_cr  = theory_cr + lab_cr

        # ── Theory grade ──────────────────────────────────────────
        oof100    = bool(c.get("out_of_100", False))
        theory_pct = compute_total_percent(
            mid       = c.get("midterm_marks",   0),
            sessional = c.get("sessional_marks",  0),
            final     = c.get("final_marks",      0),
            out_of_100 = oof100,
            total_100  = c.get("total_100_marks", 0),
        )
        theory_grade, theory_gp, theory_remark = grade_from_percent(theory_pct)

        # ── Lab grade (1 credit hour, out of 100) ─────────────────
        lab_pct   = 0.0
        lab_grade = "—"
        lab_gp    = 0.0
        lab_remark = ""
        if has_lab and lab_cr > 0:
            lab_marks = float(c.get("lab_marks", 0) or 0)
            lab_pct   = round(max(0.0, min(100.0, lab_marks)), 2)
            lab_grade_letter, lab_gp, lab_remark = grade_from_percent(lab_pct)
            lab_grade = lab_grade_letter

        # ── Combined weighted GPA contribution ────────────────────
        if has_lab and lab_cr > 0:
            combined_gp = (theory_gp * theory_cr + lab_gp * lab_cr) / total_cr
        else:
            combined_gp = theory_gp

        results.append({
            **c,
            "theory_credits":  theory_cr,
            "lab_credits":     lab_cr,
            "total_credits":   total_cr,
            "total_percent":   theory_pct,
            "grade":           theory_grade,
            "grade_point":     round(combined_gp, 2),
            "remark":          theory_remark,
            "lab_pct":         lab_pct,
            "lab_grade":       lab_grade,
            "lab_gp":          lab_gp,
            "lab_remark":      lab_remark,
        })

        sem_gp      += combined_gp * total_cr
        sem_credits += total_cr

    current_gpa  = round(sem_gp / sem_credits, 2) if sem_credits else 0.0
    overall_gp   = prev_cgpa * prev_credits + sem_gp
    overall_cr   = prev_credits + sem_credits
    overall_cgpa = round(overall_gp / overall_cr, 2) if overall_cr else current_gpa

    return jsonify({
        "results":         results,
        "current_gpa":     current_gpa,
        "overall_cgpa":    overall_cgpa,
        "sem_credits":     sem_credits,
        "overall_credits": overall_cr,
    })


@app.route("/api/predict", methods=["POST"])
def predict():
    """
    What-if prediction: given marks for one course, return predicted CGPA.
    Expects JSON: { midterm_marks, sessional_marks, final_marks,
                    out_of_100, total_100_marks, total_credits,
                    previous_cgpa, previous_credits }
    """
    d = request.get_json()

    oof100 = bool(d.get("out_of_100", False))
    pct    = compute_total_percent(
        mid        = d.get("midterm_marks",   0),
        sessional  = d.get("sessional_marks",  0),
        final      = d.get("final_marks",      0),
        out_of_100 = oof100,
        total_100  = d.get("total_100_marks",  0),
    )
    grade, gp, remark = grade_from_percent(pct)

    prev_cgpa    = float(d.get("previous_cgpa",    0))
    prev_credits = float(d.get("previous_credits", 0))
    total_cr     = float(d.get("total_credits",    3))

    overall_gp   = prev_cgpa * prev_credits + gp * total_cr
    overall_cr   = prev_credits + total_cr
    pred_cgpa    = round(overall_gp / overall_cr, 2) if overall_cr else gp

    return jsonify({
        "total_percent":  pct,
        "grade":          grade,
        "grade_point":    gp,
        "remark":         remark,
        "predicted_cgpa": pred_cgpa,
        "cgpa_change":    round(pred_cgpa - prev_cgpa, 2),
    })


@app.route("/api/export/excel", methods=["POST"])
def export_excel():
    """Export grade results as a formatted .xlsx file."""
    data = request.get_json()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grade Report"

    # Styles
    purple  = PatternFill(start_color="5C35A8", end_color="5C35A8", fill_type="solid")
    hdr_fnt = Font(bold=True, color="FFFFFF", size=11)
    thin    = Side(style="thin")
    border  = Border(left=thin, right=thin, top=thin, bottom=thin)
    center  = Alignment(horizontal="center", vertical="center")

    # Title row
    ws.merge_cells("A1:L1")
    ws["A1"]            = "Student Grade Report — Software Engineering"
    ws["A1"].font       = Font(bold=True, size=15)
    ws["A1"].alignment  = center

    # Sub-title row
    ws.merge_cells("A2:L2")
    ws["A2"]           = (f"Semester: {data.get('semester', '')}"
                          f"   |   Generated: {datetime.now().strftime('%B %d, %Y')}")
    ws["A2"].alignment = center

    # Summary row
    ws["A4"] = "Current GPA:"
    ws["B4"] = data.get("current_gpa",  0)
    ws["C4"] = "Overall CGPA:"
    ws["D4"] = data.get("overall_cgpa", 0)

    # Column headers
    headers = [
        "Course", "Teacher", "Theory Cr", "Lab Cr", "Total Cr",
        "Mid /30", "Sessional /20", "Final /50", "Theory %",
        "Lab /100", "Lab Grade", "Grade", "Grade Point", "Remark"
    ]
    for col, heading in enumerate(headers, start=1):
        cell           = ws.cell(row=6, column=col, value=heading)
        cell.fill      = purple
        cell.font      = hdr_fnt
        cell.alignment = center
        cell.border    = border

    # Data rows
    for row_idx, r in enumerate(data.get("results", []), start=7):
        oof = r.get("out_of_100", False)
        has_lab = r.get("has_lab", False)
        row_vals = [
            r.get("course_name",    ""),
            r.get("teacher",        "") or "—",
            r.get("theory_credits", ""),
            r.get("lab_credits",    "") if has_lab else "—",
            r.get("total_credits",  ""),
            "—" if oof else r.get("midterm_marks",   ""),
            "—" if oof else r.get("sessional_marks",  ""),
            "—" if oof else r.get("final_marks",      ""),
            r.get("total_percent",  ""),
            r.get("lab_marks", "—") if has_lab else "—",
            r.get("lab_grade",  "—") if has_lab else "—",
            r.get("grade",          ""),
            r.get("grade_point",    ""),
            r.get("remark",         ""),
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = border
            cell.alignment = center

    # Column widths
    col_widths = [28, 22, 9, 7, 8, 8, 13, 10, 9, 9, 10, 7, 11, 16]
    for col_letter, width in zip("ABCDEFGHIJKLMN", col_widths):
        ws.column_dimensions[col_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"grade_report_{datetime.now().strftime('%Y%m%d')}.xlsx",
    )


@app.route("/api/export/pdf", methods=["POST"])
def export_pdf():
    """Export grade results as a formatted .pdf file."""
    data = request.get_json()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        leftMargin=0.5 * inch, rightMargin=0.5 * inch,
    )
    styles = getSampleStyleSheet()
    elems  = []

    # Header
    elems.append(Paragraph("<b>Student Grade Report</b>", styles["Title"]))
    elems.append(Paragraph(
        f"Software Engineering  |  {data.get('semester', '')}",
        styles["Normal"],
    ))
    elems.append(Spacer(1, 0.15 * inch))

    # Summary table
    summary = Table(
        [[
            "Current GPA",  data.get("current_gpa",  0),
            "Overall CGPA", data.get("overall_cgpa", 0),
            "Generated",    datetime.now().strftime("%b %d, %Y"),
        ]],
        colWidths=[1.2 * inch] * 6,
    )
    summary.setStyle(TableStyle([
        ("FONTNAME",      (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 9),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elems.append(summary)
    elems.append(Spacer(1, 0.15 * inch))

    # Results table
    table_data = [["Course", "Teacher", "Cr", "Mid/30", "Ses/20", "Fin/50", "Theory%", "Lab/100", "Lab Gr", "Grade", "GP", "Remark"]]
    for r in data.get("results", []):
        has_lab   = r.get("has_lab", False)
        lab_note  = f"+{int(r.get('lab_credits', 0))}L" if has_lab else ""
        oof       = r.get("out_of_100", False)
        mid_val   = "—" if oof else r.get("midterm_marks",  "")
        ses_val   = "—" if oof else r.get("sessional_marks", "")
        fin_val   = "—" if oof else r.get("final_marks",     "")
        table_data.append([
            r.get("course_name",   ""),
            r.get("teacher",       "") or "—",
            f"{int(r.get('total_credits', 0))}{lab_note}",
            mid_val, ses_val, fin_val,
            r.get("total_percent", ""),
            r.get("lab_marks",  "—") if has_lab else "—",
            r.get("lab_grade",  "—") if has_lab else "—",
            r.get("grade",         ""),
            r.get("grade_point",   ""),
            r.get("remark",        ""),
        ])

    col_widths = [i * inch for i in [1.8, 1.2, 0.55, 0.55, 0.55, 0.55, 0.65, 0.65, 0.55, 0.55, 0.5, 1.0]]
    tbl = Table(table_data, colWidths=col_widths)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1,  0), colors.HexColor("#5C35A8")),
        ("TEXTCOLOR",     (0, 0), (-1,  0), colors.white),
        ("FONTNAME",      (0, 0), (-1,  0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 8),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("GRID",          (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#F0ECF9")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elems.append(tbl)

    doc.build(elems)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"grade_report_{datetime.now().strftime('%Y%m%d')}.pdf",
    )


if __name__ == "__main__":
    app.run(debug=True)